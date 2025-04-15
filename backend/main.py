from fastapi import FastAPI, Form, HTTPException
from fastapi.responses import StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side
from io import BytesIO
import json

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000", "https://eggfileprocessor.netlify.app"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ✅ Exact DOE blue (RGB 221, 235, 247)
doe_fill = PatternFill(
    patternType="solid",
    fgColor="DDEBF7"
)

thin_border = Border(
    left=Side(style='thin', color='FF808080'),
    right=Side(style='thin', color='FF808080'),
    top=Side(style='thin', color='FF808080'),
    bottom=Side(style='thin', color='FF808080')
)

# ✅ Left/top alignment like DOE file
left_top_align = Alignment(horizontal='left', vertical='top')

@app.post("/process/")
async def process_grades(grades: str = Form(...), absences: str = Form(None), rules: str = Form(...)):
    try:
        grades_list = [grade.strip() for grade in grades.split("\n") if grade.strip()]
    except Exception:
        raise HTTPException(status_code=400, detail="Error reading grades.")

    absences_list = []
    if absences:
        absences_list = [int(a.strip()) for a in absences.split("\n") if a.strip().isdigit()]

    if absences_list and len(absences_list) != len(grades_list):
        raise HTTPException(status_code=400, detail="Grades and absences length mismatch.")

    try:
        rules = json.loads(rules)
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="Invalid JSON format for rules.")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Processed Grades"
    sheet.append(["Original Grade", "Absences", "Updated Grade", "Comment 1", "Comment 2", "Comment 3"])

    for idx, grade in enumerate(grades_list):
        absence = absences_list[idx] if absences_list else None
        updated_grade = grade
        comments = ["", "", ""]

        for rule in rules:
            rule_comments = rule.get("comments", ["", "", ""])
            special = rule.get("specialGrade")

            if isinstance(special, dict) and special.get("value"):
                if grade.strip().upper() == special["value"].strip().upper():
                    change_to = rule.get("changeTo")
                    if change_to and change_to != "N/A":
                        updated_grade = change_to
                    comments = ["" if c == "N/A" else c for c in rule_comments]
                    break

            elif special in (None, {}, "", "N/A"):
                absence_range = rule.get("absenceRange")
                if absence_range:
                    if absence is None:
                        continue  # skip rule if absence data not provided
                    min_abs = absence_range.get("min", 0)
                    max_abs = absence_range.get("max", 46)
                    if not (min_abs <= absence <= max_abs):
                        continue

                try:
                    grade_num = float(grade)
                    min_grade = float(rule.get("minGrade", 0))
                    max_grade = float(rule.get("maxGrade", 100))
                    if min_grade <= grade_num <= max_grade:
                        absence_range = rule.get("absenceRange")
                        if absence_range:
                            min_abs = absence_range.get("min", 0)
                            max_abs = absence_range.get("max", 46)
                            if absence is None or not (min_abs <= absence <= max_abs):
                                continue
                        change_to = rule.get("changeTo")
                        if change_to and change_to != "N/A":
                            updated_grade = change_to
                        comments = ["" if c == "N/A" else c for c in rule_comments]
                        break
                except ValueError:
                    continue

        row_idx = sheet.max_row + 1

        # Original Grade (no formatting)
        sheet.cell(row=row_idx, column=1, value=grade)

        # Absences
        sheet.cell(row=row_idx, column=2, value=absence)

        # ✅ Updated Grade: format always applied
        updated_cell = sheet.cell(row=row_idx, column=3)
        updated_cell.value = str(updated_grade) if updated_grade is not None else ""
        updated_cell.number_format = "General"
        updated_cell.fill = doe_fill
        updated_cell.alignment = left_top_align
        updated_cell.border = thin_border

        # ✅ Comment 1–3: format always applied
        for i, comment in enumerate(comments):
            col_idx = 4 + i
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.value = str(comment) if comment else ""
            cell.number_format = "General"
            cell.fill = doe_fill
            cell.alignment = left_top_align
            cell.border = thin_border

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=processed_grades.xlsx"},
    )
